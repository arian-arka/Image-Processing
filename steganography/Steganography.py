import numpy
import numpy as np
from PIL import Image
import openpyxl
import json
import math
import multiprocessing
from scipy.fft import dctn
import cv2


def set_bit(v, index, x):
    mask = 1 << index
    v &= ~mask
    if x:
        v |= mask
    return v

def getWindowOfList(l, x, y, w, h):
    width, height = len(l), len(l[0])

    actualWidth = w if x + w <= width else width - x
    actualHeight = h if y + h <= height else height - y

    return {
        'window': [[l[i][j] for j in range(y, y + actualHeight)] for i in range(x, x + actualWidth)],
        'w': actualWidth,
        'h': actualHeight
    }

def getWindow(img, x, y, w, h):
    width, height = img.size

    actualWidth = w if x + w <= width else width - x
    actualHeight = h if y + h <= height else height - y

    return {
        'window': [[img.getpixel((i, j)) for j in range(y, y + actualHeight)] for i in range(x, x + actualWidth)],
        'w': actualWidth,
        'h': actualHeight
    }

def dctOfWindow(window, m, n):
    # dct = [[0 for j in range(8)] for i in range(8)]
    # mat = window['window']
    # for i in range(8):
    #     for j in range(8):
    #         ci = 1 / (8 ** 0.5) if i == 0 else (2 / 8) ** 0.5
    #         cj = 1 / (8 ** 0.5) if j == 0 else (2 / 8) ** 0.5
    #         s = 0
    #         for k in range(8):
    #             for l in range(8):
    #                 s += mat[k][l] * math.cos((2 * k + 1) * i * math.pi / (2 * 8)) * math.cos(
    #                     (2 * l + 1) * j * math.pi / (2 * 8))
    #         dct[i][j] = ci * cj * s
    # d = dctn(window['window'])
    d = cv2.dct(np.float32(window['window']))
    # print(d)
    # # print(dctn(window['window']))
    # print(window['window'])
    # exit(1)
    return 255 if d[m[0]][m[1]] > d[n[0]][n[1]] else 0

def getKeys():
    f = open('keys.json', 'r')
    data = json.loads(f.read())
    f.close()
    return data

class Steganography:
    def __init__(self, src):
        self.img = Image.open(src)

    def save(self, dst):
        self.img.save(dst)

    def findPasswordByDct(self, m, n):
        img = self.img
        width, height = img.size
        newImg = Image.new('L', (width // 8, height // 8), 'white')
        for i in range(0, width, 8):
            print(i)
            for j in range(0, height, 8):
                window = getWindow(img, i, j, 8, 8)
                bit = dctOfWindow(window, m, n)
                newImg.putpixel((i // 8, j // 8), bit)

        self.img = newImg
        return self

    def findPasswordLsb(self, xKeys, yKeys):
        img = self.img
        width, height = img.size
        newImg = Image.new('L', img.size, 'white')
        for i in range(width):
            for j in range(height):
                coords = (yKeys[j][i] - 1, xKeys[j][i] - 1)
                bit = img.getpixel(coords)[0] & 1
                newImg.putpixel((i, j), 255 if bit else 0)
        self.img = newImg
        return self

    def setPasswordLsb(self, password, xKeys, yKeys):
        img = self.img
        width, height = img.size
        newImg = Image.open(password).convert('L')
        for i in range(width):
            for j in range(height):
                pixelOfPassword = newImg.getpixel((i, j))
                coords = (yKeys[j][i] - 1, xKeys[j][i] - 1)
                pixel = img.getpixel(coords)
                r = set_bit(pixel[0], 0, 1 if pixelOfPassword else 0)
                img.putpixel(coords, (r, pixel[1], pixel[2]))
        return self

if __name__ == "__main__":
    from scipy.io import loadmat

    d = loadmat('./keyX_keyY_981531027.mat(4)(1)(1).mat')
    x = d['keyX']
    y = d['keyY']

    Steganography('./stgImageB1_981531027_R.png').setPasswordLsb('./image-1200x1600.png', x,y).save(
        'siavash.png')
    # path = './stgImage_981531027_pixelA_6_5_pixelB_5_7.png'
    # Steganography(path).findPasswordByDct((4, 5), (6, 4)).save('DCT_decrypted.png')

# keys = getKeys()
# Steganography('./stgImageB1_981531027_R.png').findPasswordLsb(keys['x'], keys['y']).save('LSB_decrypted.png')
# Steganography('./stgImageB1_981531027_R.png').setPasswordLsb('./johnF.jpg',keys['x'], keys['y']).save('LSB_encrypted.png')
# Steganography('./LSB_encrypted.png').findPasswordLsb(keys['x'], keys['y']).save('LSB_decrypted_again.png')


# sheetX = openpyxl.load_workbook('./key_x_file.xlsx')['Sheet1']
# sheetY = openpyxl.load_workbook('./key_y_file.xlsx')['Sheet1']
# data = json.dumps({
#     'x' : [[cell.value for cell in sheetX[i]] for i in range(1, sheetX.max_row + 1)],
#     'y' : [[cell.value for cell in sheetY[i]] for i in range(1, sheetY.max_row + 1)],
# })
# f= open('keys.json','w')
# f.write(data)
# f.close()
